from urllib.request import urlopen
from bs4 import BeautifulSoup
import openpyxl
from datetime import date
import time
import os
from openpyxl import Workbook

# 디렉토리 있는지 확인 후, 없으면 생성
Dir = "C:\일거래대금"
if not os.path.isdir(Dir):
	os.mkdir(Dir)

# 해당 디렉토리에 파일 생성
wb = openpyxl.Workbook()
sheet = wb.active
wb.title = '거래대금'

# 홈페이지 이동
html = 'https://coinmarketcap.com/ko/exchanges/{}'
site = ['bitmex', 'binance', 'bithumb', 'upbit']

# 결과값 리스트 형식으로 저장
List = []

# 각 사이트 별 거래대금 데이터 가져오기
for n in site:
    url = html.format(n)
    WEB = urlopen(url)
    source = BeautifulSoup(WEB, "html.parser")
    value = source.select('.h2')[0]
    List.append(value.text)

# 달러 /,(콤마) 표기 삭제
for i in range(len(List)):
    if '$' in List[i]:
        List[i] = List[i].lstrip('$')
        List[i] = List[i].replace(',', '').strip()

# 리스트 데이터 숫자로 변환
List = [int (i) for i in List]

# 환율 가져오기 & 숫자로 변환
usd_url = urlopen("https://search.naver.com/search.naver?sm=top_hty&fbm=1&ie=utf8&query=%ED%99%98%EC%9C%A8")
usd = BeautifulSoup(usd_url, "html.parser")
values = usd.find('span', {'class' : 'recite _recite result'})
values_a = values.text
values_a = values_a.replace(",", "")
values_a = values_a.replace("원", "")
values_a = int(values_a)

# 거래대금 결과 데이터 넣기
today = date.today()
now = time.localtime()

sheet['A1'] = '#'
sheet['A2'].value = site[0]
sheet['A3'].value = site[1]
sheet['A4'].value = site[2]
sheet['A5'].value = site[3]

week = ('월', '화', '수', '목', '금', '토', '일')
while True:
    if week[now.tm_wday] == '월':
        sheet.cell(row=1, column=2).value = today
        for i in range(1, 5):
            sheet.cell(row=i + 1, column=2).value = List[i - 1]*values_a
    elif week[now.tm_wday] == '화':
        sheet.cell(row=1, column=3).value = today
        for i in range(1, 5):
            sheet.cell(row=i + 1, column=3).value = List[i - 1]*values_a
    elif week[now.tm_wday] == '수':
        sheet.cell(row=1, column=4).value = today
        for i in range(1, 5):
            sheet.cell(row=i + 1, column=4).value = List[i - 1]*values_a
    elif week[now.tm_wday] == '목':
        sheet.cell(row=1, column=5).value = today
        for i in range(1, 5):
            sheet.cell(row=i + 1, column=5).value = List[i - 1]*values_a
    elif week[now.tm_wday] == '금':
        sheet.cell(row=1, column=6).value = today
        for i in range(1, 5):
            sheet.cell(row=i + 1, column=6).value = List[i - 1]*values_a
    elif week[now.tm_wday] == '토':
        sheet.cell(row=1, column=7).value = today
        for i in range(1, 5):
            sheet.cell(row=i + 1, column=7).value = List[i - 1]*values_a
    else:
        sheet.cell(row=1, column=8).value = today
        for i in range(1, 5):
            sheet.cell(row=i + 1, column=8).value = List[i - 1]*values_a
    break

# 열 너비 조절
sheet.column_dimensions['B'].width = 15
sheet.column_dimensions['C'].width = 15
sheet.column_dimensions['D'].width = 15
sheet.column_dimensions['E'].width = 15
sheet.column_dimensions['F'].width = 15
sheet.column_dimensions['G'].width = 15
sheet.column_dimensions['H'].width = 15

#엑셀 저장
wb.save("C:\일거래대금\일거래대금.xlsx")
wb.close()